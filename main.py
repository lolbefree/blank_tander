import os
import sys
import time

import openpyxl
import pyodbc
import not_for_git
from babel.dates import format_datetime


class Blank:
    def __init__(self, gsalid):
        self.file_path = os.path.abspath(os.path.dirname(__file__))
        self.gsalid = gsalid
        self.query = f"""select g.wrkordno,b.grecno,b.BILLD,v.SERIALNO,v.ADDINFO,v.LICNO,g.DISTDRIV,c.c2, g.stype 
                        from GSALS01 g
                        join GBILS01 b on b.GSALID=g.GSALID
                        join vehi v on v.VEHIID=g.VEHIID
                        left join corws01 c on c.CODAID='HFSTEERLOC' and c.c1=v.STEERING_LOCATION
                        where g.GSALID={self.gsalid}"""
        self.wb = openpyxl.load_workbook(f"{self.file_path}\\template.xlsx")

        self.server = not_for_git.db_server
        self.database = not_for_git.db_name
        self.username = not_for_git.db_user
        self.password = not_for_git.db_pw
        self.driver = '{SQL Server}'  # Driver you need to connect to the database
        self.port = '1433'
        self.cnn = pyodbc.connect(
            'DRIVER=' + self.driver + ';PORT=port;SERVER=' + self.server + ';PORT=1443;DATABASE=' + self.database + ';UID=' + self.username +
            ';PWD=' + self.password)
        self.cursor = self.cnn.cursor()

    def create_table(self):
        ws1 = self.wb[self.wb.sheetnames[0]]
        ws2 = self.wb[self.wb.sheetnames[1]]
        data = list(self.cursor.execute(self.query))[0]
        ws1["H1"] = f"{data[-1]}{data[0]} / {data[1]}"
        ws1["A3"] = format_datetime(data[2], "d MMMM Y", locale='uk_UA')
        ws1["D10"] = data[-2]
        ws2["x12"] = f"{data[3]}, {data[4]}"
        ws2["al12"] = data[-4]
        ws2["al13"] = data[-3]
        self.wb.save(f"C:\\Users\\{os.getlogin()}\\Desktop\\W{data[0]}.xlsx")
        os.system(f"C:\\Users\\{os.getlogin()}\\Desktop\\W{data[0]}.xlsx")


if __name__ == '__main__':
    args = sys.argv[1]
    Blank(f'{args}').create_table()
